# -*- coding: utf-8 -*-
"""
--------------------------------------------------
# file : Do_excel.py.py
# author : liushen
# time : 2020/12/14 0014 20:38
---------------------------------------------------
"""
from openpyxl import load_workbook
from collections import namedtuple

from scripts.Do_config import do_config
from scripts.constants import TEST_DATAS_FILE_PATH


class HandleExcel(object):   # 如果你希望其他的实例方法可以调用，那么就在实例方法中使用self
	"""
	定义处理Excel的类
	"""
	# config = Handleconfig()
	def __init__(self,filename,sheetname=None):
		self.filename = filename
		self.sheetname = sheetname
		self.wb = load_workbook(self.filename)
		# self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active   # 三元运算，为真执行左边，为假执行else右边数据
		if self.sheetname is None:   #如果没有穿sheetname这个参数，那么默认获取第一个表单
			self.ws = self.wb.active
		else:
			self.ws = self.wb[self.sheetname]
		self.sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1,values_only=True))[0]
		self.test = namedtuple("test",self.sheet_head_tuple)
		self.test_list = []   # 定义一个存放所有test命名元组对象的空列表

	def get_tests(self):
		"""
		获取所有的测试用例
		:return:存放test命名元组的列表
		"""
		for data in self.ws.iter_rows(min_row=2,max_row=5,values_only=True):
			self.test_list.append(self.test(*data))
		return self.test_list

	def get_row(self,row):
		"""
		获取某一条测试用例
		:param row:行号
		:return:一个test对象
		"""
		if isinstance(row,int) and (2 <= row <= self.ws.max_row):
			return tuple(self.ws.iter_rows(min_row=row,max_row=row,values_only=True))[0]
		else:
			print("传入的不是大于1的整数")

	def write_result(self,row,actual,result):
		"""
		将实际值与测试用例执行的结果保存到Excel中
		:param row:保存到哪一行
		:param actual:实际结果
		:param result:测试用例执行结果'Pass''Fail'
		:return:
		"""
		if isinstance(row,int) and (2 <= row <= self.ws.max_row):
			# self.ws.cell(row=row,column=6,value=actual)
			# self.ws.cell(row=row,column=self.config("excel","actual_col"),value=actual)
			self.ws.cell(row=row,column=do_config("excel","actual_col"),value=actual)
			# self.ws.cell(row=row,column=7,value=result)
			# self.ws.cell(row=row,column=HandleExcel.config("excel","result_col"),value=result)
			self.ws.cell(row=row, column=do_config("excel", "result_col"), value=result)
			self.wb.save(self.filename)

do_excel = HandleExcel(filename=do_config("file path","test_path"))


if __name__ == '__main__':
	file_name = TEST_DATAS_FILE_PATH
	one_excel = HandleExcel(filename=TEST_DATAS_FILE_PATH)
	test = one_excel.get_tests()
	print(test)

